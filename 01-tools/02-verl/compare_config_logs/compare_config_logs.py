# -*- coding: utf-8 -*-
"""从两份 verl 训练日志中提取 TaskRunner config，生成 Excel 对比表。

依赖
----
- Python 3.8+
- openpyxl: pip install openpyxl

用法
----
    python compare_config_logs.py <log_a> <log_b> [-o out.xlsx] [--name-a A] [--name-b B]
"""
import argparse
import ast
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ANSI_RE = re.compile(r"\x1b\[[0-9;]*m")
# Ray prefix may appear more than once on a corrupted line
PREFIX_RE = re.compile(r"\(TaskRunner pid=\d+\)\s?")
START_MARKER = "{'actor_rollout_ref'"
END_MARKER = "'transfer_queue'"

# pprint dict dump: keys ('foo': ...), containers, simple literals, or wrap crumbs
DICT_BODY_RE = re.compile(
    r"""^\s*(?:
        \{|
        \}+,?\s*$|
        ,\s*$|
        :\s*$|
        \[\s*(?:\]|-?\d|'|")|
        '[^']*'\s*:|
        "[^"]*"\s*:|
        '[^']*'\s*[\]\}]*\s*,?\s*$|
        "[^"]*"\s*[\]\}]*\s*,?\s*$|
        None\s*[\]\}]*\s*,?\s*$|
        True\s*[\]\}]*\s*,?\s*$|
        False\s*[\]\}]*\s*,?\s*$|
        -?\d+(?:\.\d+)?(?:e[+-]?\d+)?\s*[\]\}]*\s*,?\s*$
    )""",
    re.VERBOSE,
)

NOISE_RE = re.compile(
    r"(UserWarning|DeprecationWarning| - INFO - | - WARNING - |"
    r"WARNING:|Generating |Filtering |Loading |Setting TOKENIZERS|"
    r"Disabled critic|validate_config|Detected jtvlm|"
    r"use_critic=|warnings\.warn|main_ppo\.py:|\[INFO\])"
)


def brace_depth_delta(text: str) -> int:
    depth = 0
    i = 0
    n = len(text)
    quote = None
    while i < n:
        ch = text[i]
        if quote:
            if ch == "\\" and i + 1 < n:
                i += 2
                continue
            if ch == quote:
                quote = None
            i += 1
            continue
        if ch in ("'", '"'):
            quote = ch
            i += 1
            continue
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
        i += 1
    return depth


def extract_body_parts(raw_no_ansi: str) -> list[str]:
    """Split a line on TaskRunner prefixes and keep only dict-like bodies."""
    parts = PREFIX_RE.split(raw_no_ansi)
    bodies = []
    for part in parts:
        part = part.strip("\n")
        if not part.strip():
            continue
        if NOISE_RE.search(part):
            continue
        if START_MARKER in part or DICT_BODY_RE.match(part):
            bodies.append(part)
    return bodies


def extract_config_dict(path: Path) -> dict:
    lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
    start = None
    for i, line in enumerate(lines):
        if START_MARKER in ANSI_RE.sub("", line):
            start = i
            break
    if start is None:
        raise RuntimeError(f"Config dict not found in {path}")

    # Detect whether this log uses Ray TaskRunner line prefixes.
    has_taskrunner_prefix = any(
        "(TaskRunner pid=" in ANSI_RE.sub("", lines[i])
        for i in range(start, min(start + 20, len(lines)))
    )

    chunks: list[str] = []
    depth = 0
    started = False
    end_seen = False

    for i in range(start, len(lines)):
        raw_no_ansi = ANSI_RE.sub("", lines[i])

        if has_taskrunner_prefix:
            if "(TaskRunner pid=" not in raw_no_ansi:
                # interleaved non-TaskRunner output; keep scanning for continuation
                if started and not end_seen:
                    continue
                if started and end_seen and depth == 0:
                    break
                continue

            bodies = extract_body_parts(raw_no_ansi)
            if not bodies:
                # TaskRunner noise (warnings, info logs). Skip while incomplete.
                if started and depth > 0:
                    continue
                if started and depth == 0:
                    break
                continue
        else:
            # Plain pprint dump (e.g. NPU .out without Ray prefixes)
            body = raw_no_ansi.rstrip("\n")
            if not body.strip():
                if started and end_seen and depth == 0:
                    break
                continue
            if NOISE_RE.search(body) and START_MARKER not in body and END_MARKER not in body:
                if started and end_seen and depth == 0:
                    break
                if started and depth > 0:
                    continue
                continue
            if not (START_MARKER in body or DICT_BODY_RE.match(body) or END_MARKER in body):
                if started and end_seen and depth == 0:
                    break
                if started and depth > 0:
                    continue
                continue
            bodies = [body]

        for body in bodies:
            chunks.append(body)
            depth += brace_depth_delta(body)
            if "{" in body:
                started = True
            if END_MARKER in body:
                end_seen = True
            if started and depth == 0:
                # Space-join tolerates Ray-split pprint wraps ('key':\n None\n ,)
                joined = " ".join(chunks)
                try:
                    obj = ast.literal_eval(joined)
                except (SyntaxError, ValueError) as e:
                    preview = joined[-300:]
                    raise RuntimeError(
                        f"Failed to parse config from {path.name} near line {i + 1}: {e}\n"
                        f"...{preview!r}"
                    ) from e
                print(
                    f"{path.name}: parsed OK (end~L{i + 1}), "
                    f"top-level keys={len(obj)} -> {list(obj.keys())}"
                )
                return obj

    raise RuntimeError(
        f"Unbalanced / unparsable config dict in {path.name}, "
        f"depth={depth}, chunks={len(chunks)}, end_seen={end_seen}"
    )


def flatten(obj, prefix=""):
    items = {}
    if isinstance(obj, dict):
        if not obj:
            items[prefix] = {}
            return items
        for k, v in obj.items():
            key = f"{prefix}.{k}" if prefix else str(k)
            items.update(flatten(v, key))
    elif isinstance(obj, (list, tuple)):
        if not obj:
            items[prefix] = []
            return items
        if all(not isinstance(x, (dict, list, tuple)) for x in obj):
            items[prefix] = list(obj) if isinstance(obj, tuple) else obj
        else:
            for idx, v in enumerate(obj):
                items.update(flatten(v, f"{prefix}[{idx}]"))
    else:
        items[prefix] = obj
    return items


def value_to_str(v):
    if v is ...:
        return "<MISSING>"
    if isinstance(v, str):
        return v
    return repr(v)


def values_equal(a, b):
    if a is ... or b is ...:
        return False
    return value_to_str(a) == value_to_str(b)


def parse_args():
    parser = argparse.ArgumentParser(
        description="从两份 verl 训练日志中提取 TaskRunner config，生成 Excel 对比表。"
    )
    parser.add_argument("log_a", type=Path, help="日志 A（例如 NPU 侧 .out / .log）")
    parser.add_argument("log_b", type=Path, help="日志 B（例如 GPU 侧 .log）")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="输出 xlsx 路径；默认写入当前目录 config_compare_<name_a>_vs_<name_b>.xlsx",
    )
    parser.add_argument(
        "--name-a",
        default=None,
        help="Excel 中日志 A 的列名，默认用日志文件名（不含后缀）",
    )
    parser.add_argument(
        "--name-b",
        default=None,
        help="Excel 中日志 B 的列名，默认用日志文件名（不含后缀）",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    log_a = args.log_a.expanduser().resolve()
    log_b = args.log_b.expanduser().resolve()
    if not log_a.is_file():
        raise FileNotFoundError(f"log_a not found: {log_a}")
    if not log_b.is_file():
        raise FileNotFoundError(f"log_b not found: {log_b}")

    name_a = args.name_a or log_a.stem
    name_b = args.name_b or log_b.stem
    out_xlsx = (
        args.output.expanduser().resolve()
        if args.output
        else Path.cwd() / f"config_compare_{name_a}_vs_{name_b}.xlsx"
    )
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    cfg_a = extract_config_dict(log_a)
    cfg_b = extract_config_dict(log_b)
    flat_a = flatten(cfg_a)
    flat_b = flatten(cfg_b)

    all_keys = sorted(set(flat_a) | set(flat_b), key=lambda x: x.lower())

    title_all = "全部配置比对"
    title_diff = "差异项"
    title_sum = "汇总"
    status_same = "相同"
    status_diff = "不同"
    status_only_a = "仅" + name_a
    status_only_b = "仅" + name_b
    col_status = "状态"

    wb = Workbook()
    ws_all = wb.active
    ws_all.title = title_all
    headers = ["key", name_a, name_b, col_status]
    ws_diff = wb.create_sheet(title_diff)
    ws_sum = wb.create_sheet(title_sum)

    fill_header = PatternFill("solid", fgColor="1F4E79")
    font_header = Font(color="FFFFFF", bold=True)
    fill_diff = PatternFill("solid", fgColor="FFF2CC")
    fill_only_a = PatternFill("solid", fgColor="FCE4D6")
    fill_only_b = PatternFill("solid", fgColor="DDEBF7")
    fill_same = PatternFill("solid", fgColor="E2EFDA")

    same = diff = only_a = only_b = 0
    rows = []
    for key in all_keys:
        va = flat_a.get(key, ...)
        vb = flat_b.get(key, ...)
        if va is ...:
            status = status_only_b
            only_b += 1
            fill = fill_only_b
        elif vb is ...:
            status = status_only_a
            only_a += 1
            fill = fill_only_a
        elif values_equal(va, vb):
            status = status_same
            same += 1
            fill = fill_same
        else:
            status = status_diff
            diff += 1
            fill = fill_diff
        rows.append(([key, value_to_str(va), value_to_str(vb), status], fill, status))

    for ws in (ws_all, ws_diff):
        ws.append(headers)
        for col, _h in enumerate(headers, 1):
            cell = ws.cell(1, col)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row, fill, status in rows:
        ws_all.append(row)
        for col in range(1, 5):
            c = ws_all.cell(ws_all.max_row, col)
            c.fill = fill
            c.alignment = Alignment(vertical="top", wrap_text=True)
        if status != status_same:
            ws_diff.append(row)
            for col in range(1, 5):
                c = ws_diff.cell(ws_diff.max_row, col)
                c.fill = fill
                c.alignment = Alignment(vertical="top", wrap_text=True)

    for ws in (ws_all, ws_diff):
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 70
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 55
        ws.column_dimensions["D"].width = 14
        ws.auto_filter.ref = ws.dimensions

    ws_sum.append(["item", "value"])
    for col in (1, 2):
        c = ws_sum.cell(1, col)
        c.fill = fill_header
        c.font = font_header
    ws_sum.append(["total_keys", len(all_keys)])
    ws_sum.append(["same", same])
    ws_sum.append(["diff", diff])
    ws_sum.append([f"only_{name_a}", only_a])
    ws_sum.append([f"only_{name_b}", only_b])
    ws_sum.append(["file_A", str(log_a)])
    ws_sum.append(["file_B", str(log_b)])
    ws_sum.column_dimensions["A"].width = 20
    ws_sum.column_dimensions["B"].width = 100

    wb.save(out_xlsx)
    print(f"Wrote {out_xlsx}")
    print(f"total={len(all_keys)} same={same} diff={diff} only_a={only_a} only_b={only_b}")


if __name__ == "__main__":
    main()
